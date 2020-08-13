#encoding=utf-8
import openpyxl
from openpyxl.styles import Border, Side, Font,Fill,PatternFill,fills
import time
#version: openpyxl 3.0.4
class ParseExcel:

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color = None) # 设置字体的颜色
        # 颜色对应的RGB值
        self.RGBDict = {'red': 'FF0000', 'green': 'FF008B00'}

    def loadWorkBook(self, excelPathAndName):
        """
        :函数功能: 将excel文件加载到内存，并获取其workbook对象
        :参数:
            excelPathAndName: string, excel文件所在绝对路径
        :返回: workbook对象
        """
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook

    def getSheetByName(self, sheetName):
        """
        :函数功能: 通过sheet名获取sheet对象
        :参数:
            sheetName: string，sheet名
        :返回值: sheet object
        """
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self, sheetIndex):
        """
        :函数功能: 通过索引号获取sheet对象
        :参数:
            sheetIndex: int，sheet索引号
        :返回值: sheet object
        """
        try:
            sheetname = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self, sheet):
        """
        :函数功能: 获取sheet中存在数据区域的结束行号
        :参数:
            sheet: sheet object
        :返回值: int，数据区域的结束行号
        """
        return sheet.max_row

    def getColsNumber(self, sheet):
        """
        :函数功能: 获取sheet中存在数据区域的结束列号
        :参数:
            sheet: sheet object
        :返回值: int，数据区域的结束列号
        """
        return sheet.max_column

    def getStartRowNumber(self, sheet):
        """
        :函数功能: 获取sheet中有数据区域的起始的行号
        :参数:
            sheet: sheet object
        :返回值: int，数据区域的起始的行号
        """
        return sheet.min_row

    def getStartColNumber(self, sheet):
        """
        :函数功能: 获取sheet中有数据区域的开始的列号
        :参数:
            sheet: sheet object
        :返回值: int，数据区域的开始的列号
        """
        return sheet.min_column

    def getRow(self, sheet, rowNo):
        """
        :函数功能: 获取sheet中的行对象
        :参数:
            sheet: sheet object
            rowNo: int，行索引号, 下标从1开始，1表示第一行...
        :返回值: object，一行中所有的数据内容组成的tuple对象
        """
        try:
            return tuple(sheet.rows)[rowNo - 1]
        except Exception as e:
            raise e

    def getColumn(self, sheet, colNo):
        """
        :函数功能: 获取sheet中的列对象
        :参数:
            sheet: sheet object
            colNo: int，列索引号，下标从1开始，1表示第一列...
        :返回值: Object，一列中所有的数据内容组成tuple对象
        """
        try:
            return tuple(sheet.columns)[colNo - 1]
        except Exception as e:
            raise e

    def getCellOfValue(self, sheet, rowNo = None, colsNo = None):
        """
        :函数功能: 获取指定表格中指定单元格的值
        :参数:
            sheet: sheet object
            rowNo: int，行索引号，下标从1开始，1表示第一行...
            colsNo: int，列索引号，下标从1开始，1表示第一列...
        :返回值: string or int，指定单元格的内容
        :示例:
            getCellOfValue(sheet, rowNo = 1, colsNo = 2)
        """
        if  rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo, column = colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient params of cell !")

    def getCellOfObject(self, sheet, rowNo = None, colsNo = None):
        """
        :函数功能: 获取指定sheet中的指定单元格对象
        :参数:
            sheet: sheet object
            rowNo: int，行索引号，下标从1开始，1表示第一行...
            colsNo: int，列索引号，下标从1开始，1表示第一列...
        :返回值: object，指定单元格对象
        :示例:
            getCellObject(sheet, coordinate = 'A1')
            or
            getCellObject(sheet, rowNo = 1, colsNo = 2)
        """
        if  rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo, column = colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient params of cell !")

    def writeCell(self, sheet, content, rowNo = None, colsNo = None, fontColor = None, fillColor = None):
        """
        :函数功能: 向指定sheet中的指定单元格写入数据
        :参数:
            sheet: sheet object
            content: string/int，所写内容
            rowNo: int，行索引号，下标从1开始，1表示第一行...
            colsNo: int，列索引号，下标从1开始，1表示第一列...
            fontColor: string，所写内容颜色，red / green
            fillColor:string, 填充颜色， red/green
        :返回值: 无
        """
        if  rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row = rowNo,column = colsNo).value = content
                # sheet.cell(row=rowNo, column=colsNo,value=content)
                if fontColor:
                    sheet.cell(row = rowNo,column = colsNo).font = Font(color = self.RGBDict[fontColor])
                if fillColor:
                    sheet.cell(row = rowNo,column = colsNo).fill = PatternFill("solid", fgColor=self.RGBDict[fillColor])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient params of cell !")

    def writeCellCurrentTime(self, sheet,  rowNo = None, colsNo = None):
        """
        :函数功能: 向指定sheet的指定单元格中写入当前时间
        :参数:
            sheet: sheet object
            rowNo: int，行索引号，下标从1开始，1表示第一行...
            colsNo: int，列索引号，下标从1开始，1表示第一列...
        :返回值: 无
        """
        now = int(time.time())  #显示为时间戳
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
        if  rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row = rowNo, column = colsNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient params of cell !")

